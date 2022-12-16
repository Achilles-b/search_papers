# coding: utf-8
import requests
import pandas as pd
import json
from lxml import etree
import time
import openpyxl
import datetime

# ログ出力
import logging

# 取得結果をexcelに変換するときに使用
from openpyxl.styles import Border, Font, Side, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# excelファイルをもとにいろいろ操作するときに使用
from openpyxl import load_workbook


class SearchPapers():
    # init処理
    def __init__(self) -> None:
        """
        init処理
        ログ作成
        """
        self.logger = logging
        pass


    def main(self) -> None:
        """
        メイン処理
        """
        # 検索ワードと検索数を入力
        term = input("検索ワード\n")
        retmax = input("検索数\n")

        self.logger.debug(f"検索ワード:{term}, 検索数: {retmax}")
        
        #eSearchでpmidを取得
        pmids = self.eSearch(term, retmax)
        
        #論文の基本情報を取得し、pandasのDataFrame型として返す
        summary_df = self.eSummary(pmids)
        
        #更にアブストラクトをeFetchで取得し、pandas DataFrame型として返す
        abst_df = self.eFetch(pmids)
        
        #summaryとabstractを統合し一つのDataFrameとする
        result_df = pd.merge(summary_df, abst_df, on='pmid')

        # Excelに取得結果DataFrameの内容を出力
        filename, filenamexl = self.result_to_excel(term, result_df)

    def eSearch(self, term: str, retmax: int) -> json:
        """
        term,retmax に検索ワードと検索数を入れる。これらデータをもとにPubMedからpmidデータを取得する

        Args:
            term (str): 検索ワード
            retmax (int): 検索数

        Returns:
            json: 検索結果
        """
        URL = 'https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi?db=pubmed&retmode=json'
        option = '&retmax='+str(retmax)+'&term='+term
        query = URL + option
        response = requests.get(query)
        response_json = response.json()
        pmids = response_json['esearchresult']['idlist']
        return pmids

    def eSummary(self, pmids:json) -> pd.DataFrame:
        """
        取得したpmidsをもとにデータを取得する

        Args:
            pmids (json): pmidsのjsonファイル

        Returns:
            pd.DataFrame: pmid, Title, Author, Journal, Pubdateが含まれたデータフレーム
        """
        # URLは固定
        url = 'https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esummary.fcgi?db=pubmed&retmode=json&id='

        # pmidごとにurlを作成
        queries = [url + pmid for pmid in pmids]
        responses = {}

        # responseによってデータを取得
        for query in queries:
            response = requests.get(query)
            res_json = response.json()['result']
            responses.update(res_json)
            time.sleep(0.2) # 0.2秒ごとにsleepすることでサーバ負荷を減らす

        summaries = [{'pmid':pmid, 
                    'Title':responses[pmid]['title'], 
                    'Author':responses[pmid]['sortfirstauthor'], 
                    'Journal' : responses[pmid]['source'],
                    'Pubdate':responses[pmid]['epubdate']} for pmid in pmids]
        summary_df = pd.DataFrame(summaries)
        
        return summary_df

    def eFetch(self, pmids:json) -> pd.DataFrame:
        """
        pmidsをもとにAbstractの内容をとってくる

        Args:
            pmids (json): pmidのjsonファイル

        Returns:
            pd.DataFrame: Abstractの内容が含まれたデータフレーム
        """
        url_efetch = 'https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi?db=pubmed&retmode=xml&id='

        queries = [url_efetch + pmid for pmid in pmids]

        responses_abst = {}

        for query in queries:
            response = requests.get(query)
            root = etree.fromstring(response.content)
            pmid = root.find('.//PMID').text#pmidを抽出
            abst = root.findall('.//AbstractText')
            if abst is None:
                abst_text = ''
            else:
                abst_text = ''.join(root.xpath('//Abstract//*/text()'))
            responses_abst[pmid]=abst_text
            time.sleep(0.2)
            abst_df = pd.DataFrame.from_dict(responses_abst, orient='index')
            abst_df.index.name = 'pmid'
            abst_df.columns = ['Abstract']
        
        return abst_df  

    def result_to_excel(self, term: str, result_df:pd.DataFrame) -> str:
        """
        取得したDataFrameの結果をExcelに落とし込む

        Args:
            term (str): _description_
            result_df (pd.DataFrame): _description_

        Returns:
            str: _description_
        """
        # 結果を保存するExcelファイルの作成
        wb = openpyxl.Workbook() #ワークブックの作成
        ws = wb.active #ワークブックのアクティブになってるシートを選択
        ws.title='論文' #シートの名前を変更
        #フォントの設定
        normal_font = Font(name = "Times New Roman",sz = 10.5,b = False)
        header_font = Font(name = "Century Gothic", sz = 14, b = True, color = 'FFFFFFFF')
        #ヘッダーの塗りの設定
        header_fill = PatternFill(patternType = "solid", fgColor = "FF808080")
        #ヘッダーを中央揃えにする設定
        header_center = Alignment(horizontal='center',vertical = 'center')    
        #ヘッダーを太字にする設定
        header_border = Border(
                outline=True,
                left=Side(style='thin', color='FF000000'),
                right=Side(style='thin', color='FF000000'),
                top=Side(style='thin', color='FF000000'),
                bottom=Side(style='thin', color='FF000000')
                )
        
        # 取得結果DataFrameの中身をExcelファイルに書き込む
        for r in dataframe_to_rows(result_df, index=False, header=True):
            ws.append(r)

        for row in ws:
            for cell in row:
                cell.font = normal_font
                header_cell = ['A1', 'B1', 'C1', 'D1', 'E1', 'F1']

        for cell in header_cell:
            ws[cell].font = header_font
            ws[cell].fill = header_fill
            ws[cell].alignment = header_center
            ws[cell].border = header_border
        today = datetime.date.today()
        filename = today.strftime('%Y%m%d') + '_' + term + '_'
        filenamexl = filename + '.xlsx'
        
        # 作製結果を保存
        wb.save(filenamexl)

        return filename, wb

if __name__ == "__main__":
    logger = logging
    search_papers = SearchPapers()
    search_papers.main()

